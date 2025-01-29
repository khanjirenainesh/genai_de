from langchain_openai import AzureChatOpenAI
from sklearn.ensemble import IsolationForest
import snowflake.connector
import pandas as pd
from dotenv import load_dotenv
from pathlib import Path
import os
import time
from datetime import datetime
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine
import json

load_dotenv(override=True)

class Config:
    """Store environment variables for Snowflake and Azure OpenAI"""
    def __init__(self):
        self.env_vars = self._load_environment_variables()
        
    def _load_environment_variables(self) -> Dict[str, str]:
        """Load required environment variables"""
        required_vars = {
            "AZURE_OPENAI_ENDPOINT": os.environ.get("AZURE_OPENAI_ENDPOINT"),
            "AZURE_OPENAI_4o_DEPLOYMENT_NAME": os.environ.get("AZURE_OPENAI_4o_DEPLOYMENT_NAME"),
            "AZURE_OPENAI_API_VERSION": os.environ.get("AZURE_OPENAI_API_VERSION"),
            "AZURE_OPENAI_API_KEY": os.environ.get("AZURE_OPENAI_API_KEY"),
            "SNOWFLAKE_USER": os.environ.get("SNOWFLAKE_USER"),
            "SNOWFLAKE_PASSWORD": os.environ.get("SNOWFLAKE_PASSWORD"),
            "SNOWFLAKE_ACCOUNT": os.environ.get("SNOWFLAKE_ACCOUNT"),
            "SNOWFLAKE_WAREHOUSE": os.environ.get("SNOWFLAKE_WAREHOUSE"),
            "SNOWFLAKE_DATABASE": os.environ.get("SNOWFLAKE_DATABASE"),
            "SNOWFLAKE_SCHEMA": os.environ.get("SNOWFLAKE_SCHEMA")
        }
        
        missing_vars = [key for key, value in required_vars.items() if value is None]
        if missing_vars:
            raise ValueError(f"Missing required environment variables: {', '.join(missing_vars)}")
            
        return required_vars

class SnowflakeConnector:
    """Handle Snowflake database operations using SQLAlchemy."""
    def __init__(self, config: Config):
        self.config = config
        self.engine = self._create_engine()
        print("Snowflake connector initialized with SQLAlchemy")

    def _create_engine(self):
        try:
            connection_string = (
                f"snowflake://{self.config.env_vars['SNOWFLAKE_USER']}:"
                f"{self.config.env_vars['SNOWFLAKE_PASSWORD']}@"
                f"{self.config.env_vars['SNOWFLAKE_ACCOUNT']}/"
                f"{self.config.env_vars['SNOWFLAKE_DATABASE']}/"
                f"{self.config.env_vars['SNOWFLAKE_SCHEMA']}?warehouse="
                f"{self.config.env_vars['SNOWFLAKE_WAREHOUSE']}"
            )
            return create_engine(connection_string)
        except Exception as e:
            print(f"Error creating SQLAlchemy engine: {str(e)}")
            raise

    def get_table_metadata(self, table: str = None) -> pd.DataFrame:
        try:
            table_condition = f"and TABLE_NAME = '{table}'" if table else ""
            query = f"""
                SELECT 
                    TABLE_NAME, COLUMN_NAME, DATA_TYPE, IS_NULLABLE, CHARACTER_MAXIMUM_LENGTH
                FROM {self.config.env_vars['SNOWFLAKE_DATABASE']}.INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA = '{self.config.env_vars['SNOWFLAKE_SCHEMA']}' {table_condition}
            """
            metadata = pd.read_sql(query, self.engine)
            
            if metadata.empty:
                raise ValueError("The metadata query returned no results. Check the database, schema, and table parameters.")
            
            metadata.columns = [col.lower() for col in metadata.columns]  # Normalize column names to lowercase
            return metadata
        except Exception as e:
            print(f"Error fetching Snowflake metadata: {str(e)}")
            raise

    def get_table_data(self, table: str) -> pd.DataFrame:
        try:
            query = f"SELECT * FROM {self.config.env_vars['SNOWFLAKE_DATABASE']}.{self.config.env_vars['SNOWFLAKE_SCHEMA']}.{table}"
            return pd.read_sql(query, self.engine)
        except Exception as e:
            print(f"Error fetching Snowflake table data: {str(e)}")
            raise

    def close(self):
        if self.engine:
            self.engine.dispose()
            print("Snowflake SQLAlchemy engine disposed")


class AnomalyDetector:
    
    def __init__(self):
        self.anomalous_records = None
    """Detect anomalies in data using Isolation Forest"""
    def detect_anomalies(self, df: pd.DataFrame, table_name: str = "Unnamed Table") -> str:
        # make a copy of original dataframe
        df_copy = df.copy()

        # Convert all column names to strings
        df_copy.columns = df_copy.columns.astype(str)

        numeric_columns = []
        text_columns = []

        # Iterate through the columns to check if they are numeric or non-numeric
        for column in df_copy.columns:
        # Check if the column is not datetime and all values can be converted to numeric
            if not pd.api.types.is_datetime64_any_dtype(df_copy[column]) and pd.to_numeric(df_copy[column], errors='coerce').notna().all():
                numeric_columns.append(column)
            else:
                text_columns.append(column)

        if numeric_columns:
            try:
                anomaly_data = df[numeric_columns]
                
                anomaly_data_array = anomaly_data.to_numpy()
                
                # Ensure the data is numeric and reset index
                # anomaly_data = anomaly_data.apply(pd.to_numeric, errors='coerce')
                # anomaly_data = anomaly_data.dropna()
                
                model = IsolationForest(
                    contamination=0.01,
                    max_features=min(1.0, 10 / len(numeric_columns)),
                    max_samples=min(1.0, 1000 / len(anomaly_data_array)),
                    n_estimators=100,
                    random_state=42
                )

                # Fit and predict anomalies
                anomalies = model.fit_predict(anomaly_data_array)
                anomaly_indices = anomaly_data.index[anomalies == -1]

                if len(anomaly_indices) > 0:
                    self.anomalous_records = df_copy.loc[anomaly_indices]
                    return (
                        f"Detected anomalies in {len(self.anomalous_records)} rows in table '{table_name}'.\n"
                        f"Anomalous rows:\n{self.anomalous_records.to_string(index=False)}"
                    )
            except Exception as e:
                print(f"Error in anomaly detection: {str(e)}")

        return f"No anomalies detected in table '{table_name}'."

class InsightGenerator:
    """Generate insights using Azure OpenAI"""
    def __init__(self, config: Config):
        self.model = AzureChatOpenAI(
            azure_endpoint=config.env_vars["AZURE_OPENAI_ENDPOINT"],
            azure_deployment=config.env_vars["AZURE_OPENAI_4o_DEPLOYMENT_NAME"],
            openai_api_version=config.env_vars["AZURE_OPENAI_API_VERSION"],
            openai_api_key=config.env_vars["AZURE_OPENAI_API_KEY"],
        )
        self.system_prompts = {
            'anomaly': """You are a specialized data analyst expert in Snowflake databases, 
            anomaly detection, and data compliance. Your responses should be:
            1. Precise and technically accurate
            2. Focused on actionable solutions
            4. Compliant with data security standards
            5. Optimized for Snowflake's specific SQL syntax
            
            When analyzing anomalies:
            - Provide clear, implementable solutions
            - Generate efficient Snowflake SQL queries
            - Identify sensitive data requiring protection
            - Suggest appropriate masking/encryption techniques
            - Dont use json formating for sensitive data compliance suggestions.
            
            Always structure your responses in clean, parseable JSON,
            without any additional text or explanations outside the JSON structure.""",
            
            'semantic': """You are an expert data analyst specializing in semantic analysis 
            of database schemas and data quality. Your focus is on:
            1. Analyzing data type consistency
            2. Validating semantic meaning of columns
            3. Identifying data quality issues
            4. Suggesting schema optimizations
            
            When performing semantic analysis:
            - Check alignment between column names and data content
            - Identify mismatched data types
            - Suggest Snowflake-specific optimizations
            - Highlight semantic inconsistencies
            
            Provide clear, structured analysis focusing on actionable findings."""
        }
        print("Insight generator initialized")

    def generate_insights(self, prompt: str, analysis_type: str ) -> str:
        print("Generating insights from prompt")
        system_prompt = self.system_prompts.get(
            analysis_type
            # default to anomaly if type not found
        )
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": prompt}
        ]
        response = self.model.invoke(messages)
        return response.content

    def create_anomaly_prompt(self, issues: str) -> str:
        return f"""The following issues were detected in the Snowflake database:\n\n{issues}\n
                Give specific solution based on the issues.
                Don't add any extra line other than solution to the issue.
                Ensure that the following steps are applied to every table and every column.
                Don't mix up solution for different tables.
                Ensure the format intact for every table same.
                Provide specific issue with wrong values.
                
                Give solution in concise way.
                Also generate SQL query which is strictly Snowflake friendly to get issues.
                
                Also 
                1. Highlight columns that should be masked or encrypted, with compliance standards as PII, HIPAA, GDPR, SOC2 and provide suggestions
                2. Suggest appropriate masking techniques for each sensitive field. Dont use json formating for sensitive data compliance suggestions.
                
                I have the following unstructured text output containing:
                    1.Solutions to issues
                    2.SQL query
                    3.Sensitive data compliance suggestions
                 Convert this text into a proper JSON format with the following columns:
                    
                    [Issue_solution,SQL_query,Sensitive_Data_Compliance_Suggestions]
                    
                Ensure that:
                Each key in the JSON corresponds to the specified columns.
                All text is properly formatted and any unnecessary line breaks or inconsistent spacing is removed.
                If any section is missing in the input, leave the corresponding value as an empty string ("").
                Dont provide any extra texts before and after json data.
                Also Dont add any suggestion ot explaination before and after json data, Output should start with curly braces as json format.
                provide all accordingle json format so tat I ccan fetch details of specific column.
                
                """

    def create_semantic_prompt(self, data: pd.DataFrame, schema_details: pd.DataFrame, table_name: str) -> str:
        return f""" 
                Analysis for Snowflake table:
                Consider Snowflake-specific data types and variant/array columns.
                
                Sample data: 
                {data}
                
                metadata: 
                {schema_details}

                1. Scan through the records of each column to check if the data aligns with its semantic meaning.
                2. Highlight errors ONLY IF the semantic meaning does not align with the column name.
                3. Skip the columns where the semantic meaning and the data it holds is valid.
                4. Check for Snowflake-specific data type optimizations.
                5. ONLY provide column names and its issues.
                6. Go through all the columns and all the tables.
                7. Ensure the format intact.
                8. Please provide details of columns which has issues.
                """
 
class ExcelReportGenerator:
    """Generate Excel reports for analysis results"""
    def __init__(self, database: str, schema: str):
        self.database = database
        self.schema = schema
        self.workbook = Workbook()
        
        # Setup styles first
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Initialize workbook after styles are set
        self._initialize_workbook()
    
    def _initialize_workbook(self):
        # Remove default sheet and create required sheets
        self.workbook.remove(self.workbook.active)
        self.sheets = {
            'summary': self.workbook.create_sheet("Executive Summary"),
            'anomaly': self.workbook.create_sheet("Issue Analysis"),
            'semantic': self.workbook.create_sheet("Semantic Analysis")
        }
        
        # Initialize headers
        headers = {
            'summary': ["Database", "Schema", "Total Tables", "Tables Processed", 
                       "Start Time", "End Time", "Total Processing Time (s)", "Status"],
            'anomaly': ["Table Name", "Processing Time", "Total Records", 
                        "Issue solution", "SQL Query", "Sensitive Data Compliance Suggestions"],
            'semantic': ["Table Name", "Column Name", "Data Type", "Issue Type", 
                        "Issue Description", "Recommended Action"]
        }
        
        for sheet_name, header_list in headers.items():
            self._write_headers(self.sheets[sheet_name], header_list)

    def _write_headers(self, sheet, headers):
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
            sheet.column_dimensions[get_column_letter(col)].width = max(len(header) + 5, 15)

    def add_entry(self, sheet_name: str, data: List):
        sheet = self.sheets[sheet_name]
        next_row = sheet.max_row + 1
        for col, value in enumerate(data, 1):
            cell = sheet.cell(row=next_row, column=col)
            cell.value = value
            cell.border = self.border
            cell.alignment = Alignment(horizontal='left')

    def save(self, filepath: str):
        try:
            self.workbook.save(filepath)
            print(f"Report saved successfully to {filepath}")
        except Exception as e:
            print(f"Error saving report: {str(e)}")

def main():
    try:
        # Initialize components
        config = Config()
        db_connector = SnowflakeConnector(config)
        anomaly_detector = AnomalyDetector()
        insight_generator = InsightGenerator(config)
        
        # Set up reporting
        reports_dir = Path(__file__).parent.parent / "logs" / "snowflake_data_identification_reports"
        reports_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = reports_dir / f"snowflake_data_report_{timestamp}.xlsx"
        
        report_generator = ExcelReportGenerator(
            config.env_vars['SNOWFLAKE_DATABASE'],
            config.env_vars['SNOWFLAKE_SCHEMA']
        )
        
        # Process tables
        start_time = datetime.now()
        metadata = db_connector.get_table_metadata()
        print("Metadata columns:", metadata.columns)
        total_tables = len(metadata['table_name'].unique())

        tables_processed = 0
    
        for table in metadata['table_name'].unique():
            print(f"\nProcessing table: {table}")
            table_start = time.time()
            anomalous_records_count = 0
            try:
                # Get table data and metadata
                df = db_connector.get_table_data(table)
                table_metadata = db_connector.get_table_metadata(table)
                
                # Process in chunks
                chunk_size = 5000
                for i in range(0, len(df), chunk_size):
                    chunk = df[i:i + chunk_size]
                    
                    # Detect anomalies and generate insights
                    anomaly_result = anomaly_detector.detect_anomalies(chunk, table)
                    anomalous_records_count += len(anomaly_detector.anomalous_records) 
                    if "Detected" in anomaly_result:
                        anomaly_insights = insight_generator.generate_insights(
                            insight_generator.create_anomaly_prompt(anomaly_result),analysis_type='anomaly'
                        ).replace("plaintext", "").replace("json", "").replace("```", "").strip() 

                        anomaly_insights_json = json.loads(anomaly_insights)
                        
                        semantic_insights = insight_generator.generate_insights(
                            insight_generator.create_semantic_prompt(chunk, table_metadata, table),analysis_type='semantic'
                        ).replace("plaintext", "").replace("json", "").replace("```", "").strip() 
                        
                        
                        
                        # Add to report
                        processing_time = time.time() - table_start
                        report_generator.add_entry('anomaly', [
                            table, f"{processing_time:.2f}", len(df),
                            str(anomaly_insights_json.get("Issue_solution","")),
                            # str(anomaly_insights_json),
                            str(anomaly_insights_json.get("SQL_query","")),
                            str(anomaly_insights_json.get("Sensitive_Data_Compliance_Suggestions",""))
                        ])
                        
                        report_generator.add_entry('semantic', [
                            table, "", "", "Semantic Analysis",
                            semantic_insights, ""
                        ])
                
                tables_processed += 1
                
            except Exception as e:
                print(f"Error processing table {table}: {str(e)}")
                continue
        
        # Update summary
        end_time = datetime.now()
        total_time = (end_time - start_time).total_seconds()
        
        report_generator.add_entry('summary', [
            config.env_vars['SNOWFLAKE_DATABASE'],
            config.env_vars['SNOWFLAKE_SCHEMA'],
            total_tables,
            tables_processed,
            start_time.strftime('%Y-%m-%d %H:%M:%S'),
            end_time.strftime('%Y-%m-%d %H:%M:%S'),
            f"{total_time:.2f}",
            "Completed"
        ])
        
        # Save report
        report_generator.save(str(report_path))
        print(f"\nAnalysis completed. Results written to {report_path}")
        print(f"Total processing time: {total_time:.2f} seconds")
        print(f"Tables processed: {tables_processed}/{total_tables}")
        
    except Exception as e:
        print(f"Error in main execution: {str(e)}")
        raise
    finally:
        if 'db_connector' in locals():
            db_connector.close()

if __name__ == "__main__":
    main()